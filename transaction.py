import openpyxl
import leosLib


# Here is where all the necessary excel documents are opened
transactionWorkbook = openpyxl.load_workbook("transaction-history.xlsx")
transactionSheet = transactionWorkbook["Transactions"]
balanceSheet = transactionWorkbook["Inventory"]

currentCell = transactionSheet.cell(row = 100, column = 100)
currentCellTwo = balanceSheet.cell(row=1, column=1)

def fillCell(row, collum, string):
    current_cell = transactionSheet.cell(row=row, column=collum)
    if(string != None):
        current_cell.value = str(string)
    else:
        current_cell.value = None

def fillCellTwo(row, collum, string):
    current_cell = balanceSheet.cell(row=row, column=collum)
    if(string != None):
        current_cell.value = str(string)
    else:
        current_cell.value = None

def readCellTwo(row, collum):
    current_cell = balanceSheet.cell(row=int(row), column=int(collum))
    if(current_cell.value != None):
        return str(current_cell.value)
    else:
        return None

def findFirstEmpty(checkedCollum):
    currentRow = 1
    current_cell = transactionSheet.cell(row=1, column=int(checkedCollum))
    while(current_cell.value != None):
        currentRow += 1
        current_cell = transactionSheet.cell(row = int(currentRow), column = int(checkedCollum))
    return currentRow

def findFirstEmptyTwo(checkedCollum):
    currentRow = 1
    current_cell = balanceSheet.cell(row=1, column=int(checkedCollum))
    while(current_cell.value != None):
        currentRow += 1
        current_cell = balanceSheet.cell(row = int(currentRow), column = int(checkedCollum))
    return currentRow

def reorganizeBalance():
    whitespaceLocale = findFirstEmptyTwo(3)
    if(readCellTwo(whitespaceLocale + 1, 3) != None):
        currentRow = whitespaceLocale + 1
        while(readCellTwo(currentRow, 3) != None):
           for x in range(3, 7):
               tempString = readCellTwo(currentRow, x)
               fillCellTwo(currentRow - 1, x, tempString)
           currentRow += 1

        whitespaceLocale = findFirstEmptyTwo(3)
        whitespaceLocale -= 1
        for x in range(3, 7):
            fillCellTwo(whitespaceLocale, x, None)
    else:
        pass



validIn = False
usrIn = ""
main_actions_list = ["Buy an item (add it to the stock database)",
                     "Sell an item (remove it from the stock database)",
                     "Edit an entry of an item in the stock database (this is for if items are removed without being sold or if you entered a value wrong/need to edit TT or paid value of an entry",
                     "Add inventory (add an item to your trading inventory, this is for of you got items from a non trading source and want the system to consider them)"]

# Welcome message
print("Welcome to your Entropia trade manager, from here you will be prompted to enter which action you wish for the program to do, hit enter to continue")
input("")

# The main loop
while(1 == 1):

    print("Please select the action from the following list you wish to perform by entering the number to its left")
    leosLib.printList(main_actions_list, True, 1)
    while(validIn == False):
        usrIn = input("> ")
        if(leosLib.intable(usrIn) == True):
            if(int(usrIn) > len(main_actions_list) or int(usrIn) < 1):
                validIn = False
                print("That input was not a number")
            else:
                validIn = True
        else:
            print("That input was not a number")

    if(int(usrIn) == 1):
        targetRow = findFirstEmpty(3)
        targetRowTwo = findFirstEmptyTwo(3)
        transactionId = findFirstEmpty(15)
        transactionId -= 1

        print("You have selected the buy an item option, this requires you to first enter the name of the item")
        item = input("> ")

        print("Now please enter the amount of the item you purchased, this must be a whole number")
        amount = 0
        while(amount == 0):
            amount = input("> ")
            if(leosLib.intable(amount) == False):
                amount = 0
                print("The entered value was not valid")
            elif(int(amount) < 1):
                print("The entered value was not valid")

        print("Now please enter the total TT value of the item(s) purchased, this does not need to be a whole number")
        print("If the tt value is less than 0.01 ped, simply enter 0 or 0.01")
        ttValue = input("> ")

        print("Now please enter the total paid price of the item(s) purchased in ped, this also does not need to be a whole number")
        paidTotal = input("> ")

        # This part writes to the transaction sheet at the target row in the buy section
        fillCell(targetRow, 3, transactionId)

        fillCell(targetRow, 4, item)

        fillCell(targetRow, 5, amount)

        fillCell(targetRow, 6, paidTotal)

        fillCell(targetRow, 7, ttValue)

        # This part writes to the Inventory sheet at the inventory target

        fillCellTwo(targetRowTwo, 3, item)

        fillCellTwo(targetRowTwo, 4, amount)

        fillCellTwo(targetRowTwo, 5, ttValue)

        fillCellTwo(targetRowTwo, 6, paidTotal)



        # This part writes to the transaction sheet at the target row in the transaction history section

        fillCell(transactionId + 1, 15, transactionId)

        fillCell(transactionId + 1, 16, item)

        fillCell(transactionId + 1, 17, amount)

        fillCell(transactionId + 1, 18, paidTotal)

        fillCell(transactionId + 1, 19, ttValue)

        fillCell(transactionId + 1, 20, "Buy")

        transactionWorkbook.save("transaction-history.xlsx")

        print("Transaction successfully documented")


    elif(int(usrIn) == 2):
        inventoryItemList = []
        inventoryItemAmount = []
        inventoryItemValue = []
        inventoryItemPaidPed = []
        combinedList = []
        tempString = ""
        rowTotal = findFirstEmptyTwo(3)

        # This reads all the item names in the inventory sheet and adds them to a list
        for x in range(2, rowTotal):
            inventoryItemList.append(readCellTwo(x, 3))

        # This reads all the amounts in the inventory sheet and adds them to a list
        for x in range(2, rowTotal):
            inventoryItemAmount.append(readCellTwo(x, 4))

        #This reads all the TT values of the items and adds them to a list
        for x in range(2, rowTotal):
            inventoryItemValue.append(readCellTwo(x, 5))

        #This reads all the paid amounts of the items and adds them to a list
        for x in range(2, rowTotal):
            inventoryItemPaidPed.append(readCellTwo(x, 6))

        #This takes some of the item info and combines it into a list to be printed to the user
        for x in range(len(inventoryItemList)):
            tempString = "| Item: " + inventoryItemList[x] + "; Amount: " + inventoryItemAmount[x] + ";"
            combinedList.append(tempString)

        print("You have selected the sell option, next you will be showed a summary list of items in the inventory")

        chosenItem = 0
        sellAmount = 0
        sellPed = 0
        loop = True
        while(loop == True):
            print("To select an item type to see aditional information on it and continue the transaction enter the number to its left")
            leosLib.printList(combinedList, True, 1)
            usrIn = ""

            while(usrIn == ""):
                usrIn = input("> ")
                if(leosLib.intable(usrIn) == True):
                    usrIn = int(usrIn)
                    if(usrIn > len(inventoryItemList) or usrIn < 0):
                        usrIn = ""
                        print("That selection was an invalid range")
                else:
                    usrIn = ""

            print("You have selected " + inventoryItemList[usrIn - 1] + " here is the entry's information")
            print("Amount: " + inventoryItemAmount[usrIn - 1])
            print("Total TT value: " + inventoryItemValue[usrIn - 1] + " Ped")
            print("Paid price: " + inventoryItemPaidPed[usrIn - 1] + " Ped")
            if(float(inventoryItemAmount[usrIn - 1]) > 1):
                markUp = float(inventoryItemPaidPed[usrIn - 1]) - float(inventoryItemValue[usrIn - 1])
                markUp = markUp / float(inventoryItemValue[usrIn - 1])
                markUp *= 100.0
                markUp += 100.0
                markUp = str(markUp)
                print("Effective markup: " + markUp + "%")
            else:
                markUp = float(inventoryItemPaidPed[usrIn - 1]) - float(inventoryItemValue[usrIn - 1])
                markUp = str(markUp)
                print("Effective markup: " + markUp + " Ped\n")

            chosenItem = usrIn -1

            usrIn = " "

            print("If you would like to continue to sell the selected item or a certain amount of it hit enter or to go back to the inventory list enter E")
            while(usrIn == " "):
                usrIn = input("> ")
                if(usrIn.lower() == "e"):
                    break
                elif(usrIn.lower() == ""):
                    loop = False
                else:
                    usrIn = " "
                    print("That input was not valid, try again")

        print("You will now be prompted to fill in some information for the sale")
        usrIn = ""
        if(float(inventoryItemAmount[chosenItem]) > 1):
            print("Firstly how much of the item do you want to sell ?")
            while(usrIn == ""):
                usrIn = input("> ")
                if(leosLib.intable(usrIn) ==True):
                    if(float(usrIn) < 1 or float(usrIn) > float(inventoryItemAmount[chosenItem])):
                        usrIn = ""
                        print("The chosen amount was not valid, as a reminder there are " + inventoryItemAmount[chosenItem] + " of that item")
                else:
                    print("The entered value was not valid")
                    usrIn = ""
            sellAmount = int(usrIn)
        else:
            print("There is only one of the item to sell so no need to enter an amount")

        usrIn = ""
        print("Now please enter the amount in ped you are selling the item(s) for ")
        while(usrIn == ""):
            usrIn = input("> ")
            if(leosLib.intable(usrIn) == True):
                if(float(usrIn) < 0):
                    print("The entered value was not valid")
                    usrIn = ""
        sellPed = float(usrIn)

        ttSoldValue = 0
        ttSoldValue = float(inventoryItemValue[chosenItem]) / float(inventoryItemAmount[chosenItem])
        ttSoldValue *= float(sellAmount)

        paidValueLeftOver = 0
        paidValueLeftOver = float(inventoryItemPaidPed[chosenItem]) / float(inventoryItemAmount[chosenItem])
        paidValueLeftOver *= float(sellAmount)
        paidValueLeftOver = float(inventoryItemPaidPed[chosenItem]) - paidValueLeftOver

        targetRow = findFirstEmpty(3)
        targetRowTwo = findFirstEmptyTwo(3)
        transactionId = findFirstEmpty(15)
        transactionId -= 1

        # This part writes to the transaction sheet at the target row in the sell section
        fillCell(targetRow, 9, transactionId)

        fillCell(targetRow, 10, inventoryItemList[chosenItem])

        fillCell(targetRow, 11, sellAmount)

        fillCell(targetRow, 12, round(sellPed, 2))

        fillCell(targetRow, 13, round(ttSoldValue, 2))

        # This part writes to the Inventory sheet at the item being sold

        if(int(inventoryItemAmount[chosenItem]) - sellAmount > 0):
            fillCellTwo(chosenItem + 2, 3, inventoryItemList[chosenItem])

            fillCellTwo(chosenItem + 2, 4, int(inventoryItemAmount[chosenItem]) - sellAmount)

            fillCellTwo(chosenItem + 2, 5, round(float(inventoryItemValue[chosenItem]) - ttSoldValue, 2))

            fillCellTwo(chosenItem + 2, 6, round(paidValueLeftOver, 2))


        else:
            fillCellTwo(chosenItem + 2, 3, None)

            fillCellTwo(chosenItem + 2, 4, None)

            fillCellTwo(chosenItem + 2, 5, None)

            fillCellTwo(chosenItem + 2, 6, None)

            reorganizeBalance()




        # This part writes to the transaction sheet at the target row in the transaction history section

        print("Filling out transaction sheet (2/2)")

        fillCell(transactionId + 1, 15, transactionId)

        fillCell(transactionId + 1, 16, inventoryItemList[chosenItem])

        fillCell(transactionId + 1, 17, sellAmount)

        fillCell(transactionId + 1, 18, round(sellPed, 2))

        fillCell(transactionId + 1, 19, round(ttSoldValue, 2))

        fillCell(transactionId + 1, 20, "Sell")

        transactionWorkbook.save("transaction-history.xlsx")

        print("Transaction successfully documented")


    elif(int(usrIn) == 3):

        inventoryItemList = []
        inventoryItemAmount = []
        inventoryItemValue = []
        inventoryItemPaidPed = []
        combinedList = []
        tempString = ""
        rowTotal = findFirstEmptyTwo(3)
        recalculate = True
        loop = True
        while (loop == True):
            while(recalculate == True):
                # This reads all the item names in the inventory sheet and adds them to a list
                for x in range(2, rowTotal):
                    inventoryItemList.append(readCellTwo(x, 3))

                # This reads all the amounts in the inventory sheet and adds them to a list
                for x in range(2, rowTotal):
                    inventoryItemAmount.append(readCellTwo(x, 4))

                #This reads all the TT values of the items and adds them to a list
                for x in range(2, rowTotal):
                    inventoryItemValue.append(readCellTwo(x, 5))

                #This reads all the paid amounts of the items and adds them to a list
                for x in range(2, rowTotal):
                    inventoryItemPaidPed.append(readCellTwo(x, 6))

                #This takes some of the item info and combines it into a list to be printed to the user
                for x in range(len(inventoryItemList)):
                    tempString = "| Item: " + inventoryItemList[x] + ";"
                    combinedList.append(tempString)
                recalculate = False


            print("You have selected the edit an entry option, you will now be prompted with a list of inventory entrys, select your desired one by entering the number to its left, note editing an entry WILL NOT change the buy transaction it was linked too")
            sellAmount = 0
            sellPed = 0

            editLoop = True
            selectedItem = 0
            selectedEdit = 0

            leosLib.printList(combinedList, True, 1)
            usrIn = ""

            while(usrIn == ""):
                usrIn = input("> ")
                if(leosLib.intable(usrIn) == True):
                    usrIn = int(usrIn)
                    if(usrIn > len(inventoryItemList) or usrIn < 0):
                        usrIn = ""
                        print("That selection was an invalid range")
                else:
                    print("That selection was invalid")
                    usrIn = ""

                selectedItem = int(usrIn)
                usrIn = ""
                combinedListTwo = []
                combinedListTwo.append(inventoryItemList[selectedItem - 1])
                combinedListTwo.append(inventoryItemAmount[selectedItem - 1])
                combinedListTwo.append(inventoryItemValue[selectedItem - 1])
                combinedListTwo.append(inventoryItemPaidPed[selectedItem - 1])

                editList = ["Item name", "Item amount", "Total TT value", "Total paid ped"]
                while(editLoop == True):
                    print("You selected the entry for " + combinedListTwo[0] + ". Here are the edit options for that entry:")
                    print("1| Item name: " + combinedListTwo[0])
                    print("2| Item amount: " + combinedListTwo[1])
                    print("3| Total TT Value: " + combinedListTwo[2])
                    print("4| Total paid ped: " + combinedListTwo[3])
                    print("5| Delete entry")
                    print("6| Quit entry editor")
                    print("To select an option to edit enter the number to its left")
                    while(usrIn == ""):
                        usrIn = input("> ")
                        if(leosLib.intable(usrIn) != True):
                            usrIn = ""
                            print("The entered value was not valid")
                        elif(int(usrIn) > 6 or int(usrIn) < 1):
                            print("The entered value was outside of the supported range")
                            usrIn = ""
                    selectedEdit = int(usrIn)
                    usrIn = ""
                    if(selectedEdit == 6):
                        recalculate = True
                        editLoop = False
                    elif(selectedEdit != 5):
                        print("You have selected to edit the entrys " + editList[selectedEdit - 1])
                        print("The current value is " + combinedListTwo[selectedEdit - 1])
                        print("Enter the new value and hit enter to store it")
                        usrIn = input("> ")
                        combinedListTwo[selectedEdit - 1] = usrIn
                        fillCellTwo(selectedItem + 1, selectedEdit + 2, combinedListTwo[selectedEdit - 1])
                        transactionWorkbook.save("transaction-history.xlsx")
                        print("Change documented")
                        usrIn = ""
                        recalculate = True
                    else:
                        print("Are you sure you wish to delete this entry, if so hit enter otherwise enter any value to go back")
                        usrIn = input("> ")
                        if(usrIn == ""):
                            fillCellTwo(selectedItem + 1, 3, None)

                            fillCellTwo(selectedItem + 1, 4, None)

                            fillCellTwo(selectedItem + 1, 5, None)

                            fillCellTwo(selectedItem + 1, 6, None)

                            reorganizeBalance()

                            transactionWorkbook.save("transaction-history.xlsx")
                            print("Change documented")
                            recalculate = True
                            editLoop = False
                        else:
                            pass
                break
    elif(int(usrIn) == 4):
        itemName = ""
        itemAmount = 0
        itemTT = 0.0
        print("You have selected the add inventory option")
        print("Please enter the name of the item you wish to add to the inventory")
        while(usrIn == ""):
            if(usrIn == " "):
                print("That was not a valid name")
            else:
                itemName = usrIn
        print("Now please enter the amount of the item you want to document")
        while(usrIn == ""):
            if(leosLib.intable(usrIn) == False):
                usrIn = ""
                print("The entered input was invalid")
        itemAmount = int(usrIn)
        print("Now finally enter the TT value of the item, if the tt value is less than 0.01 ped enter 0. The item TT will also be written to the paid price collum for simplicity sake")
        while (usrIn == ""):
            if (usrIn == " "):
                print("That was not a valid number")
            else:
                itemTT = float(usrIn)

        targetRowTwo = findFirstEmptyTwo(3)

        fillCellTwo(targetRowTwo, 3, itemName)

        fillCellTwo(targetRowTwo, 4, itemAmount)

        fillCellTwo(targetRowTwo, 5, itemTT)

        fillCellTwo(targetRowTwo, 6, itemTT)

        transactionWorkbook.save("transaction-history.xlsx")

        print("Transaction successfully documented")

    break